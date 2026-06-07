import os
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

CSV_PATH = "daten/erfassung.csv"
XLSX_PATH = "auswertung/auswertung.xlsx"

os.makedirs("auswertung", exist_ok=True)

if os.path.exists(CSV_PATH) and os.path.getsize(CSV_PATH) > 0:
    df = pd.read_csv(CSV_PATH, sep=";", dtype=str)
else:
    df = pd.DataFrame(columns=["Datum", "Titel", "Kategorie", "Wert", "Einheit", "Notiz"])

for col in ["Datum", "Titel", "Kategorie", "Wert", "Einheit", "Notiz"]:
    if col not in df.columns:
        df[col] = ""

df["Wert_num"] = (
    df["Wert"]
    .astype(str)
    .str.replace(".", "", regex=False)
    .str.replace(",", ".", regex=False)
)
df["Wert_num"] = pd.to_numeric(df["Wert_num"], errors="coerce").fillna(0)

df["Datum_dt"] = pd.to_datetime(df["Datum"], errors="coerce")
df["Monat"] = df["Datum_dt"].dt.strftime("%Y-%m")
df["Monat"] = df["Monat"].fillna("Ohne Datum")

rohdaten = df[["Datum", "Titel", "Kategorie", "Wert", "Einheit", "Notiz"]].copy()

summen_kategorie = (
    df.groupby("Kategorie", dropna=False)["Wert_num"]
      .sum()
      .reset_index()
      .rename(columns={"Wert_num": "Summe Wert"})
      .sort_values(by="Kategorie")
)

anzahl_kategorie = (
    df.groupby("Kategorie", dropna=False)
      .size()
      .reset_index(name="Anzahl Datensätze")
      .sort_values(by="Kategorie")
)

summen_monat = (
    df.groupby("Monat", dropna=False)["Wert_num"]
      .sum()
      .reset_index()
      .rename(columns={"Wert_num": "Summe Wert"})
      .sort_values(by="Monat")
)

anzahl_monat = (
    df.groupby("Monat", dropna=False)
      .size()
      .reset_index(name="Anzahl Datensätze")
      .sort_values(by="Monat")
)

kennzahlen = pd.DataFrame([
    {"Kennzahl": "Datensätze", "Wert": len(df)},
    {"Kennzahl": "Kategorien", "Wert": int(df["Kategorie"].replace("", pd.NA).dropna().nunique())},
    {"Kennzahl": "Summe Wert", "Wert": float(df["Wert_num"].sum())},
])

with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
    rohdaten.to_excel(writer, sheet_name="Rohdaten", index=False)
    summen_kategorie.to_excel(writer, sheet_name="Summen Kategorie", index=False)
    anzahl_kategorie.to_excel(writer, sheet_name="Anzahl Kategorie", index=False)
    summen_monat.to_excel(writer, sheet_name="Summen Monat", index=False)
    anzahl_monat.to_excel(writer, sheet_name="Anzahl Monat", index=False)
    kennzahlen.to_excel(writer, sheet_name="Kennzahlen", index=False)

    wb = writer.book
    red_fill = PatternFill(fill_type="solid", start_color="E60000", end_color="E60000")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = red_fill
            cell.font = header_font
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 40)

        if ws.title in ["Summen Kategorie", "Summen Monat", "Kennzahlen"]:
            for row in range(2, ws.max_row + 1):
                if ws.title == "Kennzahlen":
                    target_cell = ws[f"B{row}"]
                else:
                    target_cell = ws[f"B{row}"]
                target_cell.number_format = '#,##0.00'

